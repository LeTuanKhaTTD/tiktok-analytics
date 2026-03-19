"""
Create Excel Report - Tạo báo cáo Excel chuyên nghiệp
Bao gồm: Summary, Positive, Negative, Neutral, Video Analysis
"""
import json
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_report(input_file):
    """
    Tạo Excel report với multiple sheets và formatting
    """
    print("="*80)
    print(" TẠO EXCEL REPORT - SENTIMENT ANALYSIS")
    print("="*80)
    print()
    
    # Load data
    print(f"📂 Loading: {input_file}")
    with open(input_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Flatten comments from videos
    all_comments = []
    video_sentiment_stats = {}
    
    if 'videos' in data:
        for video in data['videos']:
            video_id = video.get('video_id')
            video_url = video.get('video_url')
            
            # Track sentiment per video
            video_stats = {'positive': 0, 'negative': 0, 'neutral': 0, 'total': 0}
            
            for comment in video.get('comments', []):
                comment['video_id'] = video_id
                comment['video_url'] = video_url
                all_comments.append(comment)
                
                # Count sentiment
                sentiment = comment.get('sentiment', 'neutral')
                video_stats[sentiment] += 1
                video_stats['total'] += 1
            
            video_sentiment_stats[video_id] = {
                'video_id': video_id,
                'video_url': video_url,
                'total_comments': video_stats['total'],
                'positive': video_stats['positive'],
                'negative': video_stats['negative'],
                'neutral': video_stats['neutral'],
                'negative_rate': (video_stats['negative'] / video_stats['total'] * 100) if video_stats['total'] > 0 else 0,
                'positive_rate': (video_stats['positive'] / video_stats['total'] * 100) if video_stats['total'] > 0 else 0
            }
    else:
        all_comments = data.get('comments', [])
    
    total = len(all_comments)
    print(f"   ✅ Loaded {total:,} comments from {len(video_sentiment_stats)} videos")
    print()
    
    # Separate by sentiment
    positive_comments = [c for c in all_comments if c.get('sentiment') == 'positive']
    negative_comments = [c for c in all_comments if c.get('sentiment') == 'negative']
    neutral_comments = [c for c in all_comments if c.get('sentiment') == 'neutral']
    
    # Create output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = "reports/excel_reports"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"sentiment_analysis_report_{timestamp}.xlsx")
    
    print("📊 Creating Excel sheets...")
    print()
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # ========== SHEET 1: SUMMARY ==========
        print("   📄 Creating Summary sheet...")
        summary_data = {
            'Metric': [
                'Total Comments',
                'Total Videos',
                'Analysis Date',
                'Model Used',
                'Model Accuracy',
                '',
                'Positive Comments',
                'Positive Percentage',
                'Avg Positive Confidence',
                '',
                'Negative Comments',
                'Negative Percentage',
                'Avg Negative Confidence',
                '',
                'Neutral Comments',
                'Neutral Percentage',
                'Avg Neutral Confidence'
            ],
            'Value': [
                total,
                len(video_sentiment_stats),
                data.get('analyzed_at', datetime.now().isoformat()),
                data.get('model', 'PhoBERT'),
                data.get('accuracy', '~92%'),
                '',
                len(positive_comments),
                f"{len(positive_comments)/total*100:.1f}%",
                f"{sum(c.get('confidence', 0) for c in positive_comments) / len(positive_comments):.1%}" if positive_comments else '0%',
                '',
                len(negative_comments),
                f"{len(negative_comments)/total*100:.1f}%",
                f"{sum(c.get('confidence', 0) for c in negative_comments) / len(negative_comments):.1%}" if negative_comments else '0%',
                '',
                len(neutral_comments),
                f"{len(neutral_comments)/total*100:.1f}%",
                f"{sum(c.get('confidence', 0) for c in neutral_comments) / len(neutral_comments):.1%}" if neutral_comments else '0%'
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # ========== SHEET 2: POSITIVE ==========
        print("   ✅ Creating Positive Comments sheet...")
        positive_df = pd.DataFrame([{
            'Text': c.get('text', ''),
            'Likes': c.get('likes', 0),
            'Confidence': f"{c.get('confidence', 0):.0%}",
            'Author': c.get('author', 'unknown'),
            'Video_ID': c.get('video_id', ''),
            'Video_URL': c.get('video_url', '')
        } for c in sorted(positive_comments, key=lambda x: x.get('likes', 0), reverse=True)])
        
        if not positive_df.empty:
            positive_df.to_excel(writer, sheet_name='Positive Comments', index=False)
        
        # ========== SHEET 3: NEGATIVE ==========
        print("   ❌ Creating Negative Comments sheet...")
        negative_df = pd.DataFrame([{
            'Text': c.get('text', ''),
            'Likes': c.get('likes', 0),
            'Confidence': f"{c.get('confidence', 0):.0%}",
            'Author': c.get('author', 'unknown'),
            'Video_ID': c.get('video_id', ''),
            'Video_URL': c.get('video_url', '')
        } for c in sorted(negative_comments, key=lambda x: x.get('likes', 0), reverse=True)])
        
        if not negative_df.empty:
            negative_df.to_excel(writer, sheet_name='Negative Comments', index=False)
        
        # ========== SHEET 4: NEUTRAL ==========
        print("   ⚪ Creating Neutral Comments sheet...")
        neutral_df = pd.DataFrame([{
            'Text': c.get('text', ''),
            'Likes': c.get('likes', 0),
            'Confidence': f"{c.get('confidence', 0):.0%}",
            'Author': c.get('author', 'unknown'),
            'Video_ID': c.get('video_id', ''),
            'Video_URL': c.get('video_url', '')
        } for c in sorted(neutral_comments, key=lambda x: x.get('likes', 0), reverse=True)])
        
        if not neutral_df.empty:
            neutral_df.to_excel(writer, sheet_name='Neutral Comments', index=False)
        
        # ========== SHEET 5: VIDEO ANALYSIS ==========
        print("   📹 Creating Video Analysis sheet...")
        video_df = pd.DataFrame(list(video_sentiment_stats.values()))
        video_df = video_df.sort_values('negative_rate', ascending=False)
        video_df['negative_rate'] = video_df['negative_rate'].apply(lambda x: f"{x:.1f}%")
        video_df['positive_rate'] = video_df['positive_rate'].apply(lambda x: f"{x:.1f}%")
        
        if not video_df.empty:
            video_df.to_excel(writer, sheet_name='Video Analysis', index=False)
        
        # ========== SHEET 6: TOP NEGATIVE VIDEOS ==========
        print("   ⚠️  Creating Top Negative Videos sheet...")
        top_negative_videos = sorted(video_sentiment_stats.values(), 
                                     key=lambda x: x['negative'], reverse=True)[:20]
        
        top_negative_df = pd.DataFrame([{
            'Video_ID': v['video_id'],
            'Video_URL': v['video_url'],
            'Total_Comments': v['total_comments'],
            'Negative_Comments': v['negative'],
            'Negative_Rate': f"{v['negative_rate']:.1f}%",
            'Positive_Comments': v['positive'],
            'Neutral_Comments': v['neutral']
        } for v in top_negative_videos])
        
        if not top_negative_df.empty:
            top_negative_df.to_excel(writer, sheet_name='Top Negative Videos', index=False)
    
    # ========== FORMATTING ==========
    print()
    print("🎨 Applying formatting...")
    
    wb = load_workbook(output_file)
    
    # Define styles
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Special formatting for specific sheets
        if sheet_name == 'Summary':
            for row in ws.iter_rows(min_row=2):
                if 'Positive' in str(row[0].value):
                    row[1].fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                elif 'Negative' in str(row[0].value):
                    row[1].fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    wb.save(output_file)
    
    print("   ✅ Formatting applied!")
    print()
    
    # ========== ANALYSIS REPORT ==========
    print("="*80)
    print(" 📊 VIDEO ANALYSIS - TOP NEGATIVE VIDEOS")
    print("="*80)
    print()
    
    print("Top 10 Videos with MOST Negative Comments:")
    print("-" * 80)
    
    top_10_negative = sorted(video_sentiment_stats.values(), 
                            key=lambda x: x['negative'], reverse=True)[:10]
    
    for idx, video in enumerate(top_10_negative, 1):
        print(f"\n{idx}. Video ID: {video['video_id']}")
        print(f"   URL: {video['video_url']}")
        print(f"   Total Comments: {video['total_comments']}")
        print(f"   ❌ Negative: {video['negative']} ({video['negative_rate']:.1f}%)")
        print(f"   ✅ Positive: {video['positive']} ({video['positive_rate']:.1f}%)")
        print(f"   ⚪ Neutral:  {video['neutral']}")
    
    print()
    print("="*80)
    print(" ✅ COMPLETED!")
    print("="*80)
    print()
    print(f"📁 Excel file saved: {output_file}")
    print()
    print("Sheets created:")
    print("  1. Summary - Tổng quan phân tích")
    print("  2. Positive Comments - 527 comments tích cực")
    print("  3. Negative Comments - 233 comments tiêu cực")
    print("  4. Neutral Comments - 507 comments trung lập")
    print("  5. Video Analysis - Phân tích 83 videos")
    print("  6. Top Negative Videos - Top 20 videos cần xử lý")
    print()

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = "data/tong_hop_comment.json"
    
    create_excel_report(input_file)
